import openpyxl
import os

sheetname = 'Sheet1'
def arrangeExcel(path):
    name = ""
    set1 = set()
    excel = openpyxl.load_workbook(path)
    sheet = excel[sheetname]
    for row in sheet.rows:
        for cell in row:
            if cell.value != None:
                set1.add(cell.value)

    file = openpyxl.Workbook()
    sheet1 = file.active    
    sheet1.title = "test"
    path1 = "/Users/h0057/Desktop/%s.xlsx" % 'test'
    list1 = list(set1)
    for idx in range(0, len(list1)):
        sheet1.cell(row=idx+1, column=1, value= list1[idx])
    file.save(path1)
    print('success')


def outputStrings(path):
    dict = {}
    excel = openpyxl.load_workbook(path)
    sheet = excel[sheetname]
    for row in sheet.rows:
        dict[row[0].value] = row[1].value
        # print(dict)

    with open('/Users/h0057/Desktop/base.strings','w') as file:
        for k,v in dict.items():
            text = '"{0}" = "{1}\";\n'.format(k, v)
            print(text)
            file.write(text)




# output excel 
# arrangeExcel("/Users/h0057/Desktop/VT-Ios文字翻译表_FIN.xlsx")

#output .strings
outputStrings("/Users/h0057/Desktop/VT-iOS.xlsx")